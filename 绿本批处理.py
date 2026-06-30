#!/usr/bin/env python3
"""
绿本批处理工具 - Baidu OCR 机动车登记证书 VIN提取 + 校验
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, io, time, json, urllib.request, urllib.parse
from PIL import Image
from aip import AipOcr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

# ============ 百度OCR配置 ============
APP_ID = '123759398'
API_KEY = 'JClEZDwwrMi8d9B2XyjqgK5H'
SECRET_KEY = 'gPe36L2RpMqHKkomQcpjfDwWbSvM98B1'
# ===================================

# ============ VIN校验 ============
VIN_CHAR_MAP = {
    '0':0,'1':1,'2':2,'3':3,'4':4,'5':5,'6':6,'7':7,'8':8,'9':9,
    'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,'G':7,'H':8,
    'J':1,'K':2,'L':3,'M':4,'N':5,'P':7,'R':9,
    'S':2,'T':3,'U':4,'V':5,'W':6,'X':7,'Y':8,'Z':9
}
WEIGHTS = [8,7,6,5,4,3,2,10,0,9,8,7,6,5,4,3,2]
INVALID_CHARS = {'I','O','Q'}

def validate_vin(vin):
    v = vin.strip().upper()
    result = {'valid': False, 'vin': v, 'errors': []}
    if len(v) != 17:
        result['errors'].append(f'长度{len(v)}位（应为17位）')
        return result
    for c in v:
        if c in INVALID_CHARS:
            result['errors'].append(f'含非法字符"{c}"')
            return result
        if c not in VIN_CHAR_MAP:
            result['errors'].append(f'无法映射"{c}"')
            return result
    total = 0
    for i, ch in enumerate(v):
        total += VIN_CHAR_MAP[ch] * WEIGHTS[i]
    rem = total % 11
    calc = str(rem) if rem < 10 else 'X'
    actual = v[8]
    if calc == actual:
        result['valid'] = True
    else:
        result['errors'].append(f'校验位不符：计算={calc}，实际={actual}')
    return result
# ===================================


class VinBatchApp:
    def __init__(self, root):
        self.root = root
        self.root.title('绿本 VIN 批处理工具')
        self.root.geometry('850x600')
        self.root.resizable(True, True)
        
        self.folder_path = tk.StringVar()
        self.running = False
        self._stop = False
        self.validate_var = tk.BooleanVar(value=True)
        
        self._build_ui()
    
    def _build_ui(self):
        # === 顶部：文件夹选择 ===
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=tk.X)
        
        ttk.Label(top, text='PDF文件夹:').pack(side=tk.LEFT)
        self.folder_entry = ttk.Entry(top, textvariable=self.folder_path, width=60)
        self.folder_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(top, text='选择文件夹', command=self._select_folder).pack(side=tk.LEFT, padx=5)
        
        # === VIN校验开关 ===
        opt = ttk.Frame(self.root, padding=5)
        opt.pack(fill=tk.X)
        self.validate_cb = ttk.Checkbutton(opt, text='VIN校验（关闭时仅提取VIN，不校验合法性）',
                                           variable=self.validate_var)
        self.validate_cb.pack(side=tk.LEFT, padx=5)
        
        # === 中间：按钮 ===
        mid = ttk.Frame(self.root, padding=5)
        mid.pack(fill=tk.X)
        
        self.btn_batch = ttk.Button(mid, text='批处理', command=self._start_batch, width=12)
        self.btn_batch.pack(side=tk.LEFT, padx=5)
        
        self.btn_stop = ttk.Button(mid, text='停止', command=self._stop_batch, width=8, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, padx=5)
        
        self.progress = ttk.Progressbar(mid, mode='determinate', length=200)
        self.progress.pack(side=tk.LEFT, padx=20, fill=tk.X, expand=True)
        self.progress_label = ttk.Label(mid, text='')
        self.progress_label.pack(side=tk.LEFT, padx=5)
        
        # === 底部：日志 ===
        bottom = ttk.Frame(self.root, padding=5)
        bottom.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(bottom, text='运行日志:').pack(anchor=tk.W)
        
        log_frame = ttk.Frame(bottom)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, font=('Consolas', 10))
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 标记颜色
        self.log_text.tag_config('info', foreground='black')
        self.log_text.tag_config('ok', foreground='green')
        self.log_text.tag_config('fail', foreground='red')
        self.log_text.tag_config('title', foreground='blue', font=('Consolas', 10, 'bold'))
    
    def _select_folder(self):
        folder = filedialog.askdirectory(title='选择PDF文件夹')
        if folder:
            self.folder_path.set(folder)
    
    def _log(self, msg, tag='info'):
        self.log_text.insert(tk.END, msg + '\n', tag)
        self.log_text.see(tk.END)
        self.root.update()
    
    def _start_batch(self):
        folder = self.folder_path.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror('错误', '请先选择有效的文件夹')
            return
        
        self.running = True
        self._stop = False
        self.btn_batch.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.progress['value'] = 0
        self.progress_label['text'] = '处理中...'
        self.log_text.delete(1.0, tk.END)
        
        # 后台线程
        t = threading.Thread(target=self._batch_process, args=(folder,), daemon=True)
        t.start()
    
    def _stop_batch(self):
        self._stop = True
        self._log('用户请求中断...', 'fail')
    
    def _batch_process(self, folder):
        start_time = time.time()
        
        try:
            client = AipOcr(APP_ID, API_KEY, SECRET_KEY)
            client._timeout = 60
        except Exception as e:
            self._log(f'初始化百度OCR失败: {e}', 'fail')
            self._finish_batch()
            return
        
        # 收集PDF文件
        pdf_files = sorted([f for f in os.listdir(folder)
                           if f.lower().endswith('.pdf')
                           and os.path.isfile(os.path.join(folder, f))])
        
        if not pdf_files:
            self._log('未找到PDF文件', 'fail')
            self._finish_batch()
            return
        
        self._log(f'找到 {len(pdf_files)} 个PDF文件', 'title')
        self._log(f'输出目录: {folder}', 'info')
        self._log('-' * 60, 'info')
        
        results = {'total': len(pdf_files), 'ok': 0, 'fail': 0, 'details': []}
        
        self.progress['maximum'] = len(pdf_files)
        
        for idx, fname in enumerate(pdf_files):
            if self._stop:
                self._log('\n--- 已中断 ---', 'fail')
                break
            
            fpath = os.path.join(folder, fname)
            self.progress['value'] = idx + 1
            self.progress_label['text'] = f'{idx+1}/{len(pdf_files)}'
            self.root.update()
            
            file_start = time.time()
            
            try:
                # 用 pymupdf 渲染
                import pymupdf
                doc = pymupdf.open(fpath)
                vin_found = None
                vin_valid = None
                
                for page_idx in range(doc.page_count):
                    page = doc[page_idx]
                    pix = page.get_pixmap(dpi=300)
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    buf = io.BytesIO()
                    img.save(buf, format='JPEG', quality=90)
                    
                    r = client.vehicle_registration_certificate(buf.getvalue())
                    
                    if 'error_code' in r:
                        self._log(f'  [{fname}] OCR失败: {r.get("error_msg","")}', 'fail')
                        break
                    
                    words_result = r.get('words_result', {})
                    vin_found = words_result.get('vin', {}).get('words', '').strip()
                    
                    if vin_found:
                        break
                
                doc.close()
                
                elapsed = round(time.time() - file_start, 2)
                
                if vin_found:
                    if self.validate_var.get():
                        v_result = validate_vin(vin_found)
                        status = '通过' if v_result['valid'] else '失败'
                        tag = 'ok' if v_result['valid'] else 'fail'
                        reason = ''
                        if not v_result['valid'] and v_result['errors']:
                            reason = f' ({"; ".join(v_result["errors"])})'
                    else:
                        v_result = {'valid': True, 'errors': []}
                        status = '已提取'
                        tag = 'ok'
                        reason = '（未校验）'
                    self._log(f'  [{fname}] VIN={vin_found} {status}{reason} [{elapsed}s]', tag)
                    
                    results['details'].append({
                        'file': fname, 'vin': vin_found, 'valid': v_result['valid'],
                        'errors': v_result['errors'], 'time': elapsed
                    })
                    if v_result['valid']:
                        results['ok'] += 1
                    else:
                        results['fail'] += 1
                    
                    # 无论校验是否通过，都用VIN后六位重命名PDF
                    new_name = vin_found[-6:] + '.pdf'
                    src = fpath
                    dst = os.path.join(folder, new_name)
                    if os.path.exists(dst):
                        n = 1
                        while True:
                            new_name = f"{vin_found[-6:]}_{n}.pdf"
                            dst = os.path.join(folder, new_name)
                            if not os.path.exists(dst):
                                break
                            n += 1
                    shutil.move(src, dst)
                    self._log(f'  -> 已重命名为: {new_name}', 'ok')
                    # 更新details中的文件名
                    results['details'][-1]['file'] = new_name
                else:
                    self._log(f'  [{fname}] 未识别到VIN [{elapsed}s]', 'fail')
                    results['details'].append({
                        'file': fname, 'vin': '', 'valid': False,
                        'errors': ['未识别到VIN'], 'time': elapsed
                    })
                    results['fail'] += 1
            
            except Exception as e:
                elapsed = round(time.time() - file_start, 2)
                self._log(f'  [{fname}] 错误: {e} [{elapsed}s]', 'fail')
                results['details'].append({
                    'file': fname, 'vin': '', 'valid': False,
                    'errors': [str(e)], 'time': elapsed
                })
                results['fail'] += 1
        
        # 写结果文件
        total_elapsed = round(time.time() - start_time, 2)

        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'VIN批处理结果'
        
        # 样式定义
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=12, color='FFFFFF')
        ok_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        fail_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 标题行
        title_row = ['文件名', 'VIN码', 'VIN后六位', '校验结果', '备注', '耗时(秒)']
        for col, h in enumerate(title_row, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 数据行
        for i, d in enumerate(results['details'], 2):
            status = '有效' if d['valid'] else '无效'
            reason = '; '.join(d['errors']) if d['errors'] else ''
            vin_last6 = d['vin'][-6:] if len(d['vin']) >= 6 else d['vin']
            ws.cell(row=i, column=1, value=d['file']).border = thin_border
            ws.cell(row=i, column=2, value=d['vin']).border = thin_border
            ws.cell(row=i, column=3, value=vin_last6).border = thin_border
            ws.cell(row=i, column=4, value=status).border = thin_border
            ws.cell(row=i, column=5, value=reason).border = thin_border
            ws.cell(row=i, column=6, value=d['time']).border = thin_border
            # 颜色标记
            fill = ok_fill if d['valid'] else fail_fill
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = fill
        
        # 汇总行
        summary_row = len(results['details']) + 2
        ws.cell(row=summary_row, column=1, value='汇总').font = Font(bold=True)
        ws.cell(row=summary_row, column=1).border = thin_border
        ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=6)
        summary_text = f'总计: {results["total"]}, 有效: {results["ok"]}, 无效: {results["fail"]}  处理时间: {time.strftime("%Y-%m-%d %H:%M:%S")}  总耗时: {total_elapsed}s'
        cell = ws.cell(row=summary_row, column=2, value=summary_text)
        cell.font = Font(bold=True)
        cell.border = thin_border
        for col in range(3, 7):
            ws.cell(row=summary_row, column=col).border = thin_border
        
        # 列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        
        xlsx_path = os.path.join(folder, 'VIN结果汇总.xlsx')
        wb.save(xlsx_path)
        
        # 移动失败PDF到"失败"文件夹
        fail_dir = os.path.join(folder, '失败')
        if results['fail'] > 0:
            os.makedirs(fail_dir, exist_ok=True)
            moved = 0
            for d in results['details']:
                if not d['valid'] and d['file'].lower().endswith('.pdf'):
                    src = os.path.join(folder, d['file'])
                    dst = os.path.join(fail_dir, d['file'])
                    if os.path.exists(src):
                        shutil.move(src, dst)
                        moved += 1
            if moved > 0:
                self._log(f'已将 {moved} 个失败PDF移至: {fail_dir}', 'info')
        
        self._log('-' * 60, 'title')
        self._log(f'处理完成! 总计: {results["total"]}, 有效: {results["ok"]}, 无效: {results["fail"]}', 'title')
        self._log(f'总耗时: {total_elapsed}s', 'info')
        self._log(f'Excel结果文件: {xlsx_path}', 'info')
        if results['fail'] > 0 and os.path.exists(fail_dir):
            self._log(f'失败PDF已移至: {fail_dir}', 'info')
        
        # 保存JSON详情
        json_path = os.path.join(folder, 'VIN结果汇总.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self._finish_batch()
    
    def _finish_batch(self):
        self.running = False
        self._stop = False
        self.btn_batch.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.progress_label['text'] = ''


if __name__ == '__main__':
    root = tk.Tk()
    app = VinBatchApp(root)
    root.mainloop()
